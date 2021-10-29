import investpy
import pandas as pd
import datetime as dt
import openpyxl
from datetime import date
from openpyxl import Workbook
from openpyxl.chart import (
    LineChart,
    Reference,
)
from openpyxl.chart.axis import DateAxis

#Define the countries that you want to obtain government yields for

countries = ['U.S. 10Y','U.K. 10Y','China 10Y','Germany 10Y','Japan 10Y','Italy 10Y','Australia 10Y','Austria 10Y','Belgium 10Y',
             'Brazil 10Y','Canada 10Y','Chile 10Y','Croatia 10Y','Czech Republic 10Y',
             'France 10Y','Greece 10Y','Hong Kong 10Y','Hungary 10Y','India 10Y',
             'Indonesia 10Y','Ireland 10Y','Mexico 10Y','Netherlands 10Y',
             'New Zealand 10Y','Poland 10Y','Portugal 10Y','Russia 10Y','South Africa 10Y','South Korea 10Y',
             'Spain 10Y','Switzerland 10Y']
             
#format date and create Excel workbook

today = dt.datetime.today().strftime("%d/%m/%Y")
writer = pd.ExcelWriter('10y_Yields.xlsx')

#For each country we want to save the data to a separate tab in the Excel spreadsheet.

for country in countries:
    data = investpy.get_bond_historical_data(bond=country, from_date='01/01/1950', to_date=today)
    data_frame = pd.DataFrame(data=data)
    #data_frame = data_frame.rename(columns={'Close': country})
    data_frame.to_excel(writer,sheet_name=country)
    
writer.save()

#Now we want to create a summary page that contains all the yields so we can easily compare them. 

wb = openpyxl.load_workbook('10y_Yields.xlsx')
ws = wb.create_sheet('Summary', 0)
US10y = wb.get_sheet_by_name('U.S. 10Y')

for i in range(1,US10y.max_row):
    ws.cell(i,1).value = US10y.cell(i,1).value

for i in range(1,len(wb.sheetnames[1:])):
    for j in range(2,US10y.max_row):
        ws.cell(j,i+1).value = "=INDEX('" + wb.sheetnames[i] + "'!E:E,MATCH(Summary!A" + str(j) + ",'" + wb.sheetnames[i] + "'!A:A,0))"

for i in range(1,len(wb.sheetnames[1:])):
    ws.cell(1,i+1).value = wb.sheetnames[i]

ws = wb.get_sheet_by_name('Summary')

#create a graph on a new tab with the yields from the main economies

c1 = LineChart()
c1.title = "Bond Yields"
c1.style = 13
c1.y_axis.title = 'Yield'
c1.x_axis.title = 'Date'

data = Reference(ws, min_col=2, min_row=1, max_col=7, max_row=ws.max_row)
c1.add_data(data, titles_from_data=True)

# Style the lines
# Feel free you play around with these

s1 = c1.series[0]
s1.smooth = True # Make the line smooth
s1.graphicalProperties.line.width = 20050
s1.graphicalProperties.line.solidFill = "FF3300" # red

s2 = c1.series[1]
s2.smooth = True # Make the line smooth
s2.graphicalProperties.line.width = 20050
s2.graphicalProperties.line.solidFill = "FCF305" # yellow

s3 = c1.series[2]
s3.smooth = True # Make the line smooth
s3.graphicalProperties.line.width = 20050
s3.graphicalProperties.line.solidFill = "1FB714" # green

s4 = c1.series[3]
s4.smooth = True # Make the line smooth
s4.graphicalProperties.line.width = 20050
s4.graphicalProperties.line.solidFill = "22AAAA"

s5 = c1.series[4]
s5.smooth = True # Make the line smooth
s5.graphicalProperties.line.width = 20050
s5.graphicalProperties.line.solidFill = "33CCCC"

s6 = c1.series[5]
s6.smooth = True # Make the line smooth
s6.graphicalProperties.line.width = 20050
s6.graphicalProperties.line.solidFill = "44BBBB"

dates = Reference(ws, min_col=1, min_row=2, max_row=ws.max_row)
c1.set_categories(dates)

c1.height = 14
c1.width = 40

ws = wb.create_sheet('Graphs', 1)
ws.add_chart(c1, "A1")

#Close the workbook
wb.save('10y_Yields.xlsx')
writer.close()
openpyxl.Workbook.close('10y_Yields.xlsx')
